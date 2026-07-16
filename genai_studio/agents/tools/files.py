"""``read_file`` / ``write_file`` / ``edit_file`` — workspace-confined coding tools.

Built by a factory bound to one :class:`WorkspaceConfig` (the confinement floor lives
in ``resolve_in_workspace``). Writes are atomic (``mkstemp`` in the target dir +
``os.replace``) so a crash never leaves a half-written file. ``edit_file`` requires the
``old`` string to occur EXACTLY ONCE (mirrors the harness Edit), so the model must add
context rather than blindly replacing.
"""

from __future__ import annotations

import os
import tempfile
from pathlib import Path

from genai_studio.agents import ToolResult, tool

from ._workspace import PathEscape, WorkspaceConfig, resolve_in_workspace


def make_file_tools(ws: WorkspaceConfig) -> list:
    """Return ``[read_file, write_file, edit_file]`` bound to ``ws``."""

    @tool
    def read_file(path: str, max_bytes: int = 100_000) -> ToolResult:
        """Read a UTF-8 text file, or extract the text of a PDF, Word (.docx/.doc),
        or Excel (.xlsx/.xls) document, from the workspace.

        Args:
            path: File path, relative to the workspace root (or absolute inside it).
            max_bytes: Maximum number of bytes to read; the rest is truncated.
        """
        try:
            p = resolve_in_workspace(ws, path, for_write=False)
        except PathEscape as e:
            return ToolResult(content="", error=str(e))
        if not p.is_file():
            return ToolResult(content="", error=f"not a file: {path}")
        mb = max(0, max_bytes)
        kind = _detect_kind(p)
        if kind is not None:
            return _EXTRACTORS[kind](p, path, mb)
        try:
            size = p.stat().st_size
            data = p.read_bytes()[: max(0, max_bytes)]
            text = data.decode("utf-8")
        except UnicodeDecodeError:
            return ToolResult(content="", error=f"{path} is not UTF-8 text")
        except OSError as e:
            return ToolResult(content="", error=f"cannot read {path}: {e}")
        suffix = f"\n... (truncated at {len(data)} of {size} bytes)" if size > len(data) else ""
        return ToolResult(content=text + suffix, data={"bytes": len(data), "path": str(p)})

    @tool
    def write_file(path: str, content: str) -> ToolResult:
        """Create or overwrite a text file in the workspace (atomic write).

        Args:
            path: File path to write — must be inside the workspace and not under .git.
            content: The full new file contents.
        """
        try:
            p = resolve_in_workspace(ws, path, for_write=True)
        except PathEscape as e:
            return ToolResult(content="", error=str(e))
        existed = p.exists()
        try:
            p.parent.mkdir(parents=True, exist_ok=True)
            _atomic_write(p, content)
        except OSError as e:
            return ToolResult(content="", error=f"cannot write {path}: {e}")
        verb = "overwrote" if existed else "created"
        return ToolResult(content=f"{verb} {path} ({len(content.encode('utf-8'))} bytes)")

    @tool
    def edit_file(path: str, old: str, new: str) -> ToolResult:
        """Replace an exact string in a file. The ``old`` string must occur exactly once.

        Args:
            path: File to edit (inside the workspace).
            old: Exact existing text to replace; must be UNIQUE in the file.
            new: Replacement text.
        """
        try:
            p = resolve_in_workspace(ws, path, for_write=True)
        except PathEscape as e:
            return ToolResult(content="", error=str(e))
        if not p.is_file():
            return ToolResult(content="", error=f"not a file: {path}")
        try:
            text = p.read_text("utf-8")
        except (UnicodeDecodeError, OSError) as e:
            return ToolResult(content="", error=f"cannot read {path}: {e}")
        n = text.count(old)
        if n == 0:
            return ToolResult(content="", error="old string not found in file")
        if n > 1:
            return ToolResult(content="", error=f"old string is not unique ({n} matches); add surrounding context")
        try:
            _atomic_write(p, text.replace(old, new, 1))
        except OSError as e:
            return ToolResult(content="", error=f"cannot write {path}: {e}")
        return ToolResult(content=f"edited {path} (1 replacement)")

    return [read_file, write_file, edit_file]


# ── binary-document detection + extraction ───────────────────────────────────
# read_file transparently extracts text from PDF / Word / Excel so the model can
# "read" them instead of choking on raw bytes. Everything here FAILS OPEN: any
# detection error falls back to the plain-text path (which then reports a clean
# "not UTF-8 text" if it really is binary), and each extractor returns an
# actionable ToolResult.error (install hint / conversion hint) rather than raising.

_EXT_KIND = {".pdf": "pdf", ".docx": "docx", ".docm": "docx",
             ".xlsx": "xlsx", ".xlsm": "xlsx", ".doc": "doc", ".xls": "xls"}
_ZIP_MAGIC = b"PK\x03\x04"
_OLE_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _detect_kind(p: Path):
    """Classify a file as 'pdf'/'docx'/'xlsx'/'doc'/'xls', or ``None`` for the
    plain-text path. Extension first (fast, the common case); then a magic-byte
    sniff so a mis-named or extensionless document is still handled."""
    kind = _EXT_KIND.get(p.suffix.lower())
    if kind is not None:
        return kind
    try:
        with open(p, "rb") as f:
            head = f.read(8)
    except OSError:
        return None
    if head.startswith(b"%PDF-"):
        return "pdf"
    if head.startswith(_ZIP_MAGIC):          # OOXML is a zip — peek at the parts
        return _zip_kind(p)
    if head.startswith(_OLE_MAGIC):          # legacy Office is an OLE compound file
        return _ole_kind(p)
    return None


def _zip_kind(p: Path):
    try:
        import zipfile
        with zipfile.ZipFile(p) as z:
            names = set(z.namelist())
        if "word/document.xml" in names:
            return "docx"
        if "xl/workbook.xml" in names:
            return "xlsx"
    except Exception:
        pass
    return None


def _ole_kind(p: Path):
    try:
        import olefile
        if not olefile.isOleFile(str(p)):
            return None
        ole = olefile.OleFileIO(str(p))
        try:
            streams = {"/".join(s).lower() for s in ole.listdir()}
        finally:
            ole.close()
        if any("worddocument" in s for s in streams):
            return "doc"
        if any(s in ("workbook", "book") for s in streams):
            return "xls"
    except Exception:
        pass
    return None


def _cap(lines, max_bytes: int, note: str):
    """Join ``lines`` with newlines, stopping once ``max_bytes`` chars is reached;
    append ``note`` if truncated. Caps while building so a huge sheet/doc can't blow
    up memory."""
    out, total, truncated = [], 0, False
    for ln in lines:
        if total + len(ln) + 1 > max_bytes:
            out.append(ln[: max(0, max_bytes - total)])
            truncated = True
            break
        out.append(ln)
        total += len(ln) + 1
    text = "\n".join(out).strip()
    return (text + note) if truncated else text


def _fmt_cell(v) -> str:
    """Render a spreadsheet cell: whole floats as ints (7.0 -> '7'), None as ''."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _read_pdf(p: Path, path: str, max_bytes: int) -> ToolResult:
    """Extract a PDF's text with pypdf (page-delimited), truncated to ``max_bytes``.
    Fails with an actionable message — an install hint, not a "convert it first"
    refusal the model would otherwise emit on the raw bytes."""
    try:
        from pypdf import PdfReader
    except ImportError:
        return ToolResult(content="", error=(
            f"{path} is a PDF; text extraction needs pypdf. Install it with "
            "`pip install 'genai-studio-sdk[pdf]'` (or `pip install pypdf`), then read it again."))
    try:
        reader = PdfReader(str(p))
    except Exception as e:      # pypdf raises assorted errors on malformed/encrypted files
        return ToolResult(content="", error=f"cannot parse PDF {path}: {type(e).__name__}: {e}")
    if getattr(reader, "is_encrypted", False):
        try:
            reader.decrypt("")   # many PDFs are encrypted with an empty owner password
        except Exception:
            return ToolResult(content="", error=f"{path} is an encrypted PDF (password required)")
    parts, total, truncated, any_text = [], 0, False, False
    npages = len(reader.pages)
    for i, page in enumerate(reader.pages):
        try:
            page_text = page.extract_text() or ""
        except Exception:
            page_text = ""
        if page_text.strip():
            any_text = True
        chunk = f"\n--- page {i + 1} of {npages} ---\n{page_text}"
        if total + len(chunk) > max_bytes:
            parts.append(chunk[: max(0, max_bytes - total)])
            truncated = True
            break
        parts.append(chunk)
        total += len(chunk)
    if not any_text:      # only page delimiters came back -> scanned/image PDF
        return ToolResult(content="", error=(
            f"{path}: no extractable text ({npages} page(s)) — it is likely scanned images. "
            "OCR would be required."))
    text = "".join(parts).strip()
    if truncated:
        text += f"\n... (truncated at {max_bytes} chars of {npages}-page PDF)"
    return ToolResult(content=text, data={"pdf": True, "pages": npages, "path": str(p)})


_W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"


def _read_docx(p: Path, path: str, max_bytes: int) -> ToolResult:
    """Extract a .docx's text with the stdlib (no dependency): unzip
    ``word/document.xml`` and walk it, emitting a newline per paragraph, a tab per
    ``w:tab``, and a break per ``w:br``. Table cells are paragraphs, so they come
    through in document order."""
    import xml.etree.ElementTree as ET
    import zipfile
    try:
        with zipfile.ZipFile(p) as z:
            xml = z.read("word/document.xml")
        root = ET.fromstring(xml)
    except (zipfile.BadZipFile, KeyError, ET.ParseError, OSError) as e:
        return ToolResult(content="", error=f"cannot read .docx {path}: {type(e).__name__}: {e}")
    out: list = []

    def walk(el):
        for child in el:
            tag = child.tag
            if tag == _W + "t":
                out.append(child.text or "")
            elif tag == _W + "tab":
                out.append("\t")
            elif tag in (_W + "br", _W + "cr"):
                out.append("\n")
            else:
                walk(child)
                if tag == _W + "p":
                    out.append("\n")

    body = root.find(_W + "body")
    walk(body if body is not None else root)
    text = "".join(out).strip()
    if not text:
        return ToolResult(content="", error=f"{path}: no extractable text in the Word document.")
    text = _cap(text.split("\n"), max_bytes, f"\n... (truncated at {max_bytes} chars)")
    return ToolResult(content=text, data={"docx": True, "path": str(p)})


def _read_xlsx(p: Path, path: str, max_bytes: int) -> ToolResult:
    """Extract an .xlsx as tab-separated rows per sheet, using openpyxl."""
    try:
        import openpyxl
    except ImportError:
        return ToolResult(content="", error=(
            f"{path} is an .xlsx; reading it needs openpyxl. Install it with "
            "`pip install 'genai-studio-sdk[xlsx]'` (or `pip install openpyxl`), then read it again."))
    try:
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    except Exception as e:
        return ToolResult(content="", error=f"cannot read .xlsx {path}: {type(e).__name__}: {e}")
    n_sheets = len(wb.sheetnames)

    def lines():
        for ws in wb.worksheets:
            yield f"--- sheet: {ws.title} ---"
            for row in ws.iter_rows(values_only=True):
                yield "\t".join(_fmt_cell(c) for c in row)

    try:
        text = _cap(lines(), max_bytes, f"\n... (truncated at {max_bytes} chars)")
    finally:
        wb.close()
    return ToolResult(content=text, data={"xlsx": True, "sheets": n_sheets, "path": str(p)})


def _read_xls(p: Path, path: str, max_bytes: int) -> ToolResult:
    """Extract a legacy .xls as tab-separated rows per sheet, using xlrd."""
    try:
        import xlrd
    except ImportError:
        return ToolResult(content="", error=(
            f"{path} is a legacy .xls; reading it needs xlrd. Install it with "
            "`pip install 'genai-studio-sdk[xls]'` (or `pip install xlrd`), then read it again."))
    try:
        book = xlrd.open_workbook(str(p))
    except Exception as e:
        return ToolResult(content="", error=f"cannot read .xls {path}: {type(e).__name__}: {e}")

    def lines():
        for sh in book.sheets():
            yield f"--- sheet: {sh.name} ---"
            for rx in range(sh.nrows):
                yield "\t".join(_fmt_cell(v) for v in sh.row_values(rx))

    text = _cap(lines(), max_bytes, f"\n... (truncated at {max_bytes} chars)")
    return ToolResult(content=text, data={"xls": True, "sheets": book.nsheets, "path": str(p)})


def _read_doc(p: Path, path: str, max_bytes: int) -> ToolResult:
    """Extract a legacy Word .doc (OLE compound file) with olefile — parse the FIB,
    read the CLX piece table from the table stream, and decode each piece (8-bit
    cp1252 or 16-bit UTF-16LE per its ``fc`` flag). BEST-EFFORT: the .doc binary
    format is old and varied, so on any parse failure this returns an actionable
    'convert to .docx' message rather than raising or emitting garbage."""
    import struct
    try:
        import olefile
    except ImportError:
        return ToolResult(content="", error=(
            f"{path} is a legacy .doc; reading it needs olefile. Install it with "
            "`pip install 'genai-studio-sdk[doc]'` (or `pip install olefile`), then read it again."))
    convert_hint = (f"{path}: could not extract text from this legacy .doc. Convert it to "
                    ".docx (or PDF) and read that instead.")
    if not olefile.isOleFile(str(p)):
        return ToolResult(content="", error=convert_hint)
    ole = olefile.OleFileIO(str(p))
    try:
        wd = ole.openstream("WordDocument").read()
        flags = struct.unpack_from("<H", wd, 0x0A)[0]
        table_name = "1Table" if (flags & 0x0200) else "0Table"
        if not ole.exists(table_name):
            table_name = "0Table" if table_name == "1Table" else "1Table"
        ccp_text = struct.unpack_from("<i", wd, 0x4C)[0]
        fc_clx = struct.unpack_from("<I", wd, 0x01A2)[0]
        lcb_clx = struct.unpack_from("<I", wd, 0x01A6)[0]
        clx = ole.openstream(table_name).read()[fc_clx:fc_clx + lcb_clx]
        i = 0
        while i < len(clx) and clx[i] == 0x01:      # skip leading Prc (formatting) blocks
            i += 3 + struct.unpack_from("<H", clx, i + 1)[0]
        if i >= len(clx) or clx[i] != 0x02:         # expect the Pcdt marker
            return ToolResult(content="", error=convert_hint)
        lcb = struct.unpack_from("<I", clx, i + 1)[0]
        plc = clx[i + 5:i + 5 + lcb]
        n = (lcb - 4) // 12                          # (n+1) CPs of 4B + n PCDs of 8B
        cps = [struct.unpack_from("<I", plc, k * 4)[0] for k in range(n + 1)]
        chunks = []
        for k in range(n):
            fc = struct.unpack_from("<I", plc, 4 * (n + 1) + k * 8 + 2)[0]
            nchars = cps[k + 1] - cps[k]
            if fc & 0x40000000:                      # 8-bit cp1252, fc halved
                base = (fc & 0x3FFFFFFF) // 2
                chunks.append(wd[base:base + nchars].decode("cp1252", "replace"))
            else:                                    # 16-bit utf-16le
                chunks.append(wd[fc:fc + nchars * 2].decode("utf-16-le", "replace"))
    except Exception:
        return ToolResult(content="", error=convert_hint)
    finally:
        ole.close()
    raw = "".join(chunks)
    if ccp_text > 0:
        raw = raw[:ccp_text]                          # keep just the main document text
    # Word control marks -> readable text: \r ends a paragraph, \x07 delimits cells.
    for a, b in (("\r", "\n"), ("\x07", "\t"), ("\x0b", "\n"), ("\x0c", "\n"),
                 ("\x01", ""), ("\x02", ""), ("\x05", ""), ("\x08", "")):
        raw = raw.replace(a, b)
    text = _cap(raw.split("\n"), max_bytes, f"\n... (truncated at {max_bytes} chars)")
    if not text.strip():
        return ToolResult(content="", error=convert_hint)
    return ToolResult(content=text, data={"doc": True, "path": str(p)})


_EXTRACTORS = {"pdf": _read_pdf, "docx": _read_docx, "xlsx": _read_xlsx,
               "xls": _read_xls, "doc": _read_doc}


def _atomic_write(p: Path, content: str) -> None:
    """Write ``content`` to ``p`` atomically (temp in the same dir + os.replace)."""
    fd, tmp = tempfile.mkstemp(dir=str(p.parent), prefix=".tmp-")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(content)
        os.replace(tmp, p)
    finally:
        if os.path.exists(tmp):
            try:
                os.unlink(tmp)
            except OSError:
                pass
